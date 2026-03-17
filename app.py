from openpyxl import load_workbook
planilha = load_workbook('Book.xlsx')

pagina1 = planilha['Planilha1']
for linha in pagina1.iter_rows(values_only=True):
    print(linha)    
